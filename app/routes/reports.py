import io
from flask import Blueprint, render_template, send_file
from flask_login import login_required
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from app.extensions import db
from app.models import Product, Category, Department

reports_bp = Blueprint("reports", __name__)


@reports_bp.route("/")
@login_required
def index():
    total_products = db.session.scalar(db.select(func.count(Product.id))) or 0

    # Se utiliza 'quantity' en lugar de 'stock'
    low_stock_products = db.session.scalars(
        db.select(Product).where(Product.quantity <= Product.min_stock)
    ).all()

    # Valor total calculado con 'quantity'
    inventory_value = (
        db.session.scalar(db.select(func.sum(Product.quantity * Product.price))) or 0
    )

    # Resumen por categoría
    category_summary = db.session.execute(
        db.select(
            Category.name,
            func.count(Product.id).label("total_items"),
            func.sum(Product.quantity).label("total_stock"),
        )
        .outerjoin(Product)
        .group_by(Category.id, Category.name)
    ).all()

    # Resumen por departamento
    department_summary = db.session.execute(
        db.select(
            Department.name,
            func.count(Category.id).label("total_categories"),
        )
        .outerjoin(Category)
        .group_by(Department.id, Department.name)
    ).all()

    return render_template(
        "reports/index.html",
        total_products=total_products,
        low_stock_products=low_stock_products,
        inventory_value=inventory_value,
        category_summary=category_summary,
        department_summary=department_summary,
    )


@reports_bp.route("/exportar/excel")
@login_required
def export_excel():
    products = db.session.scalars(db.select(Product)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Encabezados
    headers = ["ID", "SKU", "Nombre", "Precio", "Stock", "Estatus"]
    ws.append(headers)

    # Estilos para el encabezado
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Filas de datos
    for p in products:
        status = "Bajo Stock" if p.is_low_stock else "Normal"
        ws.append([p.id, p.sku, p.name, float(p.price or 0), p.quantity, status])

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="reporte_inventario.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@reports_bp.route("/exportar/pdf")
@login_required
def export_pdf():
    products = db.session.scalars(db.select(Product)).all()

    output = io.BytesIO()
    pdf = canvas.Canvas(output, pagesize=letter)

    # Encabezado del documento
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, 750, "Reporte de Inventario Pro")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(50, 735, "Listado de productos registrados en el sistema")
    pdf.line(50, 725, 550, 725)

    # Contenido del reporte
    y = 700
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(50, y, "SKU")
    pdf.drawString(150, y, "Nombre")
    pdf.drawString(350, y, "Precio")
    pdf.drawString(450, y, "Stock")
    y -= 20

    pdf.setFont("Helvetica", 10)
    for p in products:
        pdf.drawString(50, y, str(p.sku))
        pdf.drawString(150, y, str(p.name)[:30])
        pdf.drawString(350, y, f"${p.price:.2f}")
        pdf.drawString(450, y, str(p.quantity))
        y -= 18

        # Control de salto de página
        if y < 50:
            pdf.showPage()
            y = 750
            pdf.setFont("Helvetica", 10)

    pdf.save()
    output.seek(0)

    return send_file(
        output,
        download_name="reporte_inventario.pdf",
        as_attachment=True,
        mimetype="application/pdf",
    )